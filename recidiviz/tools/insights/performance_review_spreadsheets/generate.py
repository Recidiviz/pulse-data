# Recidiviz - a data platform for criminal justice reform
# Copyright (C) 2024 Recidiviz, Inc.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
# =============================================================================
"""
Script for generating Excel spreadsheets with staff metrics for supervisors in ID, for use in their
performance reviews. See https://www.notion.so/recidiviz/Design-Doc-ID-2025-Performance-Reviews-1517889f4d198077b276fc1fc5569165
for details.

Usage: python -m recidiviz.tools.insights.performance_review_spreadsheets.generate
This will generate spreadsheets in the root directory of pulse-data.
"""

import logging
import os
from datetime import date
from enum import Enum, auto
from random import random

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_NUMBER
from openpyxl.worksheet.worksheet import Worksheet

from recidiviz.big_query.big_query_client import BigQueryClientImpl
from recidiviz.tools.insights.performance_review_spreadsheets.officer_aggregated_metrics import (
    OfficerAggregatedMetrics,
)
from recidiviz.tools.insights.performance_review_spreadsheets.supervisors_and_officers import (
    SupervisorsAndOfficers,
)
from recidiviz.utils.environment import GCP_PROJECT_PRODUCTION
from recidiviz.utils.metadata import local_project_id_override


class RowHeadingEnum(Enum):
    AVG_DAILY_CASELOAD = auto()
    TIMELY_RISK_ASSESSMENTS = auto()
    TIMELY_F2F_CONTACTS = auto()
    SUPERVISION_LEVEL_MISMATCH = auto()
    EARLY_DISCHARGE_GRANTS = auto()
    LSU_GRANTS = auto()
    SLD_GRANTS = auto()
    FT_DISCHARGE_GRANTS = auto()


class ColumnHeadingEnum(Enum):
    YEARLY = "All of 2024"


_ROW_HEADING_LABELS = {
    RowHeadingEnum.AVG_DAILY_CASELOAD: "# Average Daily Caseload",
    RowHeadingEnum.TIMELY_RISK_ASSESSMENTS: "Timely Risk Assessments",
    RowHeadingEnum.TIMELY_F2F_CONTACTS: "Timely F2F Contacts",
    RowHeadingEnum.SUPERVISION_LEVEL_MISMATCH: "Supervision & Risk Level Mismatch",
    RowHeadingEnum.EARLY_DISCHARGE_GRANTS: "Grants during time period",
    RowHeadingEnum.LSU_GRANTS: "Grants during time period",
    RowHeadingEnum.SLD_GRANTS: "Grants during time period",
    RowHeadingEnum.FT_DISCHARGE_GRANTS: "Grants during time period",
}

_ROW_HEADINGS = [
    "Usage",
    "# Total Logins each month",
    "# of Months in 2024 where they logged in at least once",
    "Caseload Type",
    None,
    "Outcomes Metrics",
    RowHeadingEnum.AVG_DAILY_CASELOAD,
    "# Absconsions",
    "Months flagged as having a high absconsion rate",
    "# Incarcerations",
    "Months flagged as having a high incarceration rate",
    None,
    "Early Discharge",
    "Eligible & Viewed as of end of time period",
    "Eligible & Not Viewed as of end of time period",
    "Overridden (Marked Ineligible) as of end of time period",
    RowHeadingEnum.EARLY_DISCHARGE_GRANTS,
    "Monthly grant rate",
    'Most often used "Ineligible Reason"',
    None,
    "LSU",
    "Eligible & Viewed as of end of time period",
    "Eligible & Not Viewed as of end of time period",
    "Overridden (Marked Ineligible) as of end of time period",
    RowHeadingEnum.LSU_GRANTS,
    "2024 average of monthly grant rate",
    'Most often used "Ineligible Reason"',
    None,
    "Supervision Level Downgrade",
    "Eligible & Viewed as of end of time period",
    "Eligible & Not Viewed as of end of time period",
    "Overridden (Marked Ineligible) as of end of time period",
    RowHeadingEnum.SLD_GRANTS,
    "2024 average of monthly grant rate",
    'Most often used "Ineligible Reason"',
    None,
    "Past FTRD",
    "Eligible & Viewed as of end of time period",
    "Eligible & Not Viewed as of end of time period",
    "Overridden (Marked Ineligible) as of end of time period",
    RowHeadingEnum.FT_DISCHARGE_GRANTS,
    "Monthly grant rate",
    'Most often used "Ineligible Reason"',
    None,
    "Operations Metrics",
    RowHeadingEnum.TIMELY_RISK_ASSESSMENTS,
    RowHeadingEnum.TIMELY_F2F_CONTACTS,
    RowHeadingEnum.SUPERVISION_LEVEL_MISMATCH,
]

_COLUMN_DATE_FORMAT = "%b %Y"

_COLUMN_HEADINGS = [
    None,
    ColumnHeadingEnum.YEARLY.value,
    *[date(2024, month, 1).strftime(_COLUMN_DATE_FORMAT) for month in range(1, 13)],
]

# Row numbers (1-indexed) that represent a section header
_ROW_SECTION_HEADER_INDICES = [2, 7, 14, 22, 30, 38, 46]

_GRAY_BACKGROUND = PatternFill(
    start_color="d9d9d9", end_color="d9d9d9", fill_type="solid"
)
_RED_BACKGROUND = PatternFill(
    start_color="ea9999", end_color="ea9999", fill_type="solid"
)
_BOLD_FONT = Font(bold=True)


CONDITIONAL_FORMAT_THRESHOLDS = {
    RowHeadingEnum.TIMELY_RISK_ASSESSMENTS: 0.9,
    RowHeadingEnum.TIMELY_F2F_CONTACTS: 0.9,
    RowHeadingEnum.SUPERVISION_LEVEL_MISMATCH: 0.98,
}


def create_headers(sheet: Worksheet) -> None:
    """Adds header rows + columns to the spreadsheet. This must be run before other data is added to
    the spreadsheet, because it just appends full rows."""
    data = [
        _COLUMN_HEADINGS,
        *[
            [_ROW_HEADING_LABELS[row] if isinstance(row, RowHeadingEnum) else row]
            for row in _ROW_HEADINGS
        ],
    ]
    for row in data:
        sheet.append(row)

    sheet.column_dimensions["A"].width = 48
    for index in _ROW_SECTION_HEADER_INDICES:
        sheet[f"A{index}"].font = _BOLD_FONT
        sheet[f"A{index}"].fill = _GRAY_BACKGROUND
    sheet["B1"].font = _BOLD_FONT


def get_row_index(heading: RowHeadingEnum) -> int:
    return (
        _ROW_HEADINGS.index(heading) + 2  # 1-indexed + additional 1 for empty first row
    )


def get_column_index(heading: str) -> str:
    return chr(ord("A") + _COLUMN_HEADINGS.index(heading))


def apply_conditional_formatting(sheet: Worksheet) -> None:
    for row_heading, threshold in CONDITIONAL_FORMAT_THRESHOLDS.items():
        row_idx = get_row_index(row_heading)
        col_idx = get_column_index(ColumnHeadingEnum.YEARLY.value)

        cell = f"{col_idx}{row_idx}"

        sheet.conditional_formatting.add(
            cell,
            FormulaRule(
                formula=[f"AND(NOT(ISBLANK({cell})), {cell}<{threshold})"],
                fill=_RED_BACKGROUND,
            ),
        )


def write_random_data(sheet: Worksheet) -> None:
    for row in range(2, 50):
        # We don't need to put data in the section header rows or the rows before them
        if row in _ROW_SECTION_HEADER_INDICES or row + 1 in _ROW_SECTION_HEADER_INDICES:
            continue
        for col in range(ord("B"), ord("O")):
            sheet[f"{chr(col)}{row}"] = random()


def try_set_metric_cell(
    sheet: Worksheet, metric_value: int | float | None, col_idx: str, row_idx: int
) -> None:
    if metric_value is not None:
        cell = sheet[f"{col_idx}{row_idx}"]
        cell.value = metric_value
        cell.number_format = FORMAT_NUMBER


def write_metrics(
    sheet: Worksheet,
    officer_id: str,
    officer_aggregated_metrics: OfficerAggregatedMetrics,
) -> None:
    """Write aggregated metrics to the appropriate cells in the sheet"""
    avg_daily_caseload_row = get_row_index(RowHeadingEnum.AVG_DAILY_CASELOAD)
    early_discharge_row = get_row_index(RowHeadingEnum.EARLY_DISCHARGE_GRANTS)
    lsu_row = get_row_index(RowHeadingEnum.LSU_GRANTS)
    sld_row = get_row_index(RowHeadingEnum.SLD_GRANTS)
    ft_discharge_row = get_row_index(RowHeadingEnum.FT_DISCHARGE_GRANTS)

    for metric in officer_aggregated_metrics.monthly_data[officer_id]:
        parsed_date = (metric.end_date_exclusive - relativedelta(days=1)).strftime(
            _COLUMN_DATE_FORMAT
        )
        col_idx = get_column_index(parsed_date)
        try_set_metric_cell(
            sheet, metric.avg_daily_population, col_idx, avg_daily_caseload_row
        )
        try_set_metric_cell(
            sheet, metric.task_completions_early_discharge, col_idx, early_discharge_row
        )
        try_set_metric_cell(
            sheet,
            metric.task_completions_transfer_to_limited_supervision,
            col_idx,
            lsu_row,
        )
        try_set_metric_cell(
            sheet, metric.task_completions_supervision_level_downgrade, col_idx, sld_row
        )
        try_set_metric_cell(
            sheet,
            metric.task_completions_full_term_discharge,
            col_idx,
            ft_discharge_row,
        )

    yearly_col_idx = get_column_index(ColumnHeadingEnum.YEARLY.value)
    yearly_data = officer_aggregated_metrics.yearly_data[officer_id]
    try_set_metric_cell(
        sheet, yearly_data.avg_daily_population, yearly_col_idx, avg_daily_caseload_row
    )
    try_set_metric_cell(
        sheet,
        yearly_data.task_completions_early_discharge,
        yearly_col_idx,
        early_discharge_row,
    )
    try_set_metric_cell(
        sheet,
        yearly_data.task_completions_transfer_to_limited_supervision,
        yearly_col_idx,
        lsu_row,
    )
    try_set_metric_cell(
        sheet,
        yearly_data.task_completions_supervision_level_downgrade,
        yearly_col_idx,
        sld_row,
    )
    try_set_metric_cell(
        sheet,
        yearly_data.task_completions_full_term_discharge,
        yearly_col_idx,
        ft_discharge_row,
    )


def generate_sheets() -> None:
    output_dir = os.path.join(os.path.dirname(__file__), "output")
    if not os.path.exists(output_dir):
        os.mkdir(output_dir)

    bq_client = BigQueryClientImpl()
    supervisors_to_officers = SupervisorsAndOfficers.from_bigquery(bq_client).data
    officer_aggregated_metrics = OfficerAggregatedMetrics.from_bigquery(bq_client)
    for supervisor, officers in supervisors_to_officers.items():
        wb = Workbook()
        for officer in officers:
            sheet = wb.create_sheet(title=officer.officer_name.formatted())
            create_headers(sheet)
            apply_conditional_formatting(sheet)
            write_metrics(sheet, officer.officer_id, officer_aggregated_metrics)

        # Remove the default sheet, since we created new sheets for our data.
        wb.remove(wb.worksheets[0])
        wb.save(f"{output_dir}/{supervisor.formatted()}.xlsx")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    with local_project_id_override(GCP_PROJECT_PRODUCTION):
        generate_sheets()
