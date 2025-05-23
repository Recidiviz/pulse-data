# Recidiviz - a data platform for criminal justice reform
# Copyright (C) 2024 Recidiviz, Inc.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
# =============================================================================
"""
Script for generating Excel spreadsheets with staff metrics for supervisors in ND, for use in their
performance reviews. See https://www.notion.so/recidiviz/Design-Doc-ID-2025-Performance-Reviews-1517889f4d198077b276fc1fc5569165
for details.

Usage: python -m recidiviz.tools.insights.performance_review_spreadsheets_US_ND.generate
This will generate spreadsheets in the root directory of pulse-data.
"""

import argparse
import logging
import os
import sys
from datetime import date
from enum import Enum, auto
from typing import List, Tuple

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER, FORMAT_PERCENTAGE_00
from openpyxl.worksheet.worksheet import Worksheet

from recidiviz.big_query.big_query_client import BigQueryClientImpl
from recidiviz.tools.insights.performance_review_spreadsheets_US_ND.impact_funnel_metrics import (
    ImpactFunnelMetrics,
)
from recidiviz.tools.insights.performance_review_spreadsheets_US_ND.logins import Logins
from recidiviz.tools.insights.performance_review_spreadsheets_US_ND.officer_aggregated_metrics import (
    OfficerAggregatedMetrics,
)
from recidiviz.tools.insights.performance_review_spreadsheets_US_ND.officer_aggregated_metrics_from_sandbox import (
    OfficerAggregatedMetricsFromSandbox,
)
from recidiviz.tools.insights.performance_review_spreadsheets_US_ND.snooze_metrics import (
    SnoozeMetrics,
)
from recidiviz.tools.insights.performance_review_spreadsheets_US_ND.supervisors_and_officers import (
    SupervisorsAndOfficers,
)
from recidiviz.utils.environment import GCP_PROJECT_PRODUCTION
from recidiviz.utils.metadata import local_project_id_override


class RowHeadingEnum(Enum):
    """Enum to represent each labeled row in the sheet."""

    USAGE = auto()
    NUM_LOGINS = auto()
    NUM_MONTHS_LOGGED_IN = auto()
    CASELOAD_INFO = auto()
    AVG_DAILY_CASELOAD = auto()
    CASELOAD_TYPE = auto()
    OUTCOMES_METRICS = auto()
    # ABSCONSIONS = auto()
    # INCARCERATIONS = auto()
    TIMELY_RISK_ASSESSMENTS = auto()
    TIMELY_CONTACTS = auto()
    EARLY_TERMINATION = auto()
    EARLY_TERMINATION_GRANTS = auto()
    EARLY_TERMINATION_MONTHLY_GRANT_RATE = auto()
    EARLY_TERMINATION_ELIGIBLE_VIEWED = auto()
    EARLY_TERMINATION_ELIGIBLE_NOT_VIEWED = auto()
    EARLY_TERMINATION_ELIGIBLE_NOT_VIEWED_30_DAYS = auto()
    EARLY_TERMINATION_MARKED_INELIGIBLE = auto()
    EARLY_TERMINATION_MARKED_INELIGIBLE_REASON = auto()
    OPERATIONS_METRICS = auto()


class ColumnHeadingEnum(Enum):
    YEARLY = "All of 2024"


_ROW_HEADING_LABELS = {
    RowHeadingEnum.USAGE: "Usage",
    RowHeadingEnum.NUM_LOGINS: "# Total Logins",
    RowHeadingEnum.NUM_MONTHS_LOGGED_IN: "# of Months in 2024 where they logged in at least once",
    RowHeadingEnum.CASELOAD_INFO: "Caseload Information",
    RowHeadingEnum.AVG_DAILY_CASELOAD: "# Average Daily Caseload",
    RowHeadingEnum.CASELOAD_TYPE: "Caseload Type",
    RowHeadingEnum.OUTCOMES_METRICS: "Outcomes Metrics",
    # RowHeadingEnum.ABSCONSIONS: "# Absconsions on caseload",
    # RowHeadingEnum.INCARCERATIONS: "# Re-incarcerations on caseload",
    RowHeadingEnum.EARLY_TERMINATION: "Early Termination",
    RowHeadingEnum.EARLY_TERMINATION_GRANTS: "Grants during time period",
    RowHeadingEnum.EARLY_TERMINATION_MONTHLY_GRANT_RATE: "Monthly grant rate: (# grants that month / avg daily caseload that month) or average of monthly grant rate ",
    RowHeadingEnum.EARLY_TERMINATION_ELIGIBLE_VIEWED: "Number of clients Eligible Now and Viewed as of end of time period",
    RowHeadingEnum.EARLY_TERMINATION_ELIGIBLE_NOT_VIEWED: "Number of clients Eligible and Not viewed as of end of time period",
    RowHeadingEnum.EARLY_TERMINATION_ELIGIBLE_NOT_VIEWED_30_DAYS: "Number of not-viewed clients who had been eligible for 30 or more days as of end of time period",
    RowHeadingEnum.EARLY_TERMINATION_MARKED_INELIGIBLE: "Overridden (Marked Ineligible) as of end of time period",
    RowHeadingEnum.EARLY_TERMINATION_MARKED_INELIGIBLE_REASON: 'Most often used "Ineligible Reason"',
    RowHeadingEnum.OPERATIONS_METRICS: "Operations Metrics",
    RowHeadingEnum.TIMELY_RISK_ASSESSMENTS: "Timely Risk Assessments",
    RowHeadingEnum.TIMELY_CONTACTS: "Timely F2F Contacts",
}

_ROW_HEADINGS: list[RowHeadingEnum | str | None] = [
    RowHeadingEnum.USAGE,
    RowHeadingEnum.NUM_MONTHS_LOGGED_IN,
    RowHeadingEnum.NUM_LOGINS,
    None,
    RowHeadingEnum.CASELOAD_INFO,
    RowHeadingEnum.AVG_DAILY_CASELOAD,
    RowHeadingEnum.CASELOAD_TYPE,
    # None,
    # RowHeadingEnum.OUTCOMES_METRICS,
    # RowHeadingEnum.ABSCONSIONS,
    # RowHeadingEnum.INCARCERATIONS,
    None,
    RowHeadingEnum.EARLY_TERMINATION,
    RowHeadingEnum.EARLY_TERMINATION_ELIGIBLE_VIEWED,
    RowHeadingEnum.EARLY_TERMINATION_ELIGIBLE_NOT_VIEWED,
    RowHeadingEnum.EARLY_TERMINATION_ELIGIBLE_NOT_VIEWED_30_DAYS,
    RowHeadingEnum.EARLY_TERMINATION_MARKED_INELIGIBLE,
    RowHeadingEnum.EARLY_TERMINATION_GRANTS,
    RowHeadingEnum.EARLY_TERMINATION_MONTHLY_GRANT_RATE,
    RowHeadingEnum.EARLY_TERMINATION_MARKED_INELIGIBLE_REASON,
    None,
    RowHeadingEnum.OPERATIONS_METRICS,
    RowHeadingEnum.TIMELY_RISK_ASSESSMENTS,
    RowHeadingEnum.TIMELY_CONTACTS,
]

_COLUMN_DATE_FORMAT = "%b %Y"

_COLUMN_HEADINGS = [
    None,
    ColumnHeadingEnum.YEARLY.value,
    *[date(2024, month, 1).strftime(_COLUMN_DATE_FORMAT) for month in range(1, 13)],
    *[date(2025, month, 1).strftime(_COLUMN_DATE_FORMAT) for month in range(1, 4)],
]

# Rows that represent a section header and need different formatting
_ROW_SECTION_HEADERS = [
    RowHeadingEnum.USAGE,
    RowHeadingEnum.CASELOAD_INFO,
    # RowHeadingEnum.OUTCOMES_METRICS,
    RowHeadingEnum.EARLY_TERMINATION,
    RowHeadingEnum.OPERATIONS_METRICS,
]

_GRAY_BACKGROUND = PatternFill(
    start_color="d9d9d9", end_color="d9d9d9", fill_type="solid"
)
_RED_BACKGROUND = PatternFill(
    start_color="ea9999", end_color="ea9999", fill_type="solid"
)
_GREEN_BACKGROUND = PatternFill(
    start_color="9edbbe", end_color="9edbbe", fill_type="solid"
)
_BOLD_FONT = Font(bold=True)
_BOLD_12_FONT = Font(size=12, bold=True)
_BOLD_13_FONT = Font(size=13, bold=True)

_RIGHT_BORDER = Border(right=Side(style="medium", color="999999"))
_THIN_RIGHT_BORDER = Border(right=Side(style="thin", color="999999"))
_THIN_BOTTOM_BORDER = Border(bottom=Side(style="thin", color="999999"))

NEG_CONDITIONAL_FORMAT_THRESHOLDS = {
    RowHeadingEnum.TIMELY_RISK_ASSESSMENTS: (0.8, "<"),
    RowHeadingEnum.TIMELY_CONTACTS: (0.8, "<"),
    RowHeadingEnum.EARLY_TERMINATION_ELIGIBLE_NOT_VIEWED_30_DAYS: (1, ">="),
}

POS_CONDITIONAL_FORMAT_THRESHOLDS = {
    RowHeadingEnum.TIMELY_RISK_ASSESSMENTS: (0.9, ">="),
    RowHeadingEnum.TIMELY_CONTACTS: (0.9, ">="),
}


def create_headers(
    sheet: Worksheet,
    officer_name: str,
    officer_email: str,
    officer_district: str,
    supervisor_name: str,
) -> None:
    """Adds header rows + columns to the spreadsheet. This must be run before other data is added to
    the spreadsheet, because it just appends full rows."""
    name_and_maybe_email = (
        f"{officer_name}\n{officer_email}" if officer_email else officer_name
    )
    data: list[list] = [
        [name_and_maybe_email],
        [f"LEAD OFFICER: {supervisor_name}\nREGION: {officer_district}"],
        _COLUMN_HEADINGS,
        *[
            [_ROW_HEADING_LABELS[row] if isinstance(row, RowHeadingEnum) else row]
            for row in _ROW_HEADINGS
        ],
    ]
    for index, row in enumerate(data):
        sheet.append(row)
        sheet[f"A{index+1}"].alignment = Alignment(wrap_text=True)

    sheet.column_dimensions["A"].width = 48
    sheet.column_dimensions["B"].width = 28
    for header in _ROW_SECTION_HEADERS:
        index = get_row_index(header)
        sheet[f"A{index}"].font = _BOLD_FONT
        for c in range(1, 18):
            sheet.cell(row=index, column=c).fill = _GRAY_BACKGROUND
    for cell in sheet["B3:B42"]:
        cell[0].font = _BOLD_12_FONT
        cell[0].border = _RIGHT_BORDER
    sheet["A1"].font = _BOLD_13_FONT
    sheet["A2"].font = _BOLD_FONT
    for c in range(1, 15):
        sheet.cell(row=2, column=c).border = _THIN_BOTTOM_BORDER
        sheet.cell(row=42, column=c).border = _THIN_BOTTOM_BORDER
    for cell in sheet["N3:N42"]:
        cell[0].border = _THIN_RIGHT_BORDER


def get_row_index(heading: RowHeadingEnum) -> int:
    return (
        _ROW_HEADINGS.index(heading)
        + 4  # 1-indexed + additional 3 for empty first 3 rows
    )


def get_column_index(heading: str) -> str:
    return chr(ord("A") + _COLUMN_HEADINGS.index(heading))


def apply_conditional_formatting(sheet: Worksheet) -> None:
    for row_heading, (
        threshold,
        operator,
    ) in NEG_CONDITIONAL_FORMAT_THRESHOLDS.items():
        row_idx = get_row_index(row_heading)
        col_idx = get_column_index(ColumnHeadingEnum.YEARLY.value)

        cell = f"{col_idx}{row_idx}"

        sheet.conditional_formatting.add(
            cell,
            FormulaRule(
                formula=[f"AND(NOT(ISBLANK({cell})), {cell}{operator}{threshold})"],
                fill=_RED_BACKGROUND,
            ),
        )
    for row_heading, (
        threshold,
        operator,
    ) in POS_CONDITIONAL_FORMAT_THRESHOLDS.items():
        row_idx = get_row_index(row_heading)
        col_idx = get_column_index(ColumnHeadingEnum.YEARLY.value)

        cell = f"{col_idx}{row_idx}"

        sheet.conditional_formatting.add(
            cell,
            FormulaRule(
                formula=[f"AND(NOT(ISBLANK({cell})), {cell}{operator}{threshold})"],
                fill=_GREEN_BACKGROUND,
            ),
        )


def try_set_metric_cell(
    sheet: Worksheet,
    metric_value: int | float | None,
    col_idx: str,
    row_idx: int,
    number_format: str = FORMAT_NUMBER,
) -> None:
    if metric_value is not None:
        cell = sheet[f"{col_idx}{row_idx}"]
        cell.value = metric_value
        cell.number_format = number_format


def write_metrics(
    sheet: Worksheet,
    officer_id: str,
    officer_aggregated_metrics: OfficerAggregatedMetrics,
) -> None:
    """Write aggregated metrics to the appropriate cells in the sheet"""
    avg_daily_caseload_row = get_row_index(RowHeadingEnum.AVG_DAILY_CASELOAD)
    # absconsions_row = get_row_index(RowHeadingEnum.ABSCONSIONS)
    # incarcerations_row = get_row_index(RowHeadingEnum.INCARCERATIONS)
    for metric in officer_aggregated_metrics.monthly_data[officer_id]:
        parsed_date = (metric.end_date_exclusive - relativedelta(days=1)).strftime(
            _COLUMN_DATE_FORMAT
        )
        col_idx = get_column_index(parsed_date)
        try_set_metric_cell(
            sheet, metric.avg_daily_population, col_idx, avg_daily_caseload_row
        )
        # try_set_metric_cell(sheet, metric.num_absconsions, col_idx, absconsions_row)
        # try_set_metric_cell(
        #     sheet, metric.num_incarcerations, col_idx, incarcerations_row
        # )
    yearly_col_idx = get_column_index(ColumnHeadingEnum.YEARLY.value)
    if officer_id in officer_aggregated_metrics.yearly_data:
        yearly_data = officer_aggregated_metrics.yearly_data[officer_id]
        try_set_metric_cell(
            sheet,
            yearly_data.avg_daily_population,
            yearly_col_idx,
            avg_daily_caseload_row,
        )
        # try_set_metric_cell(
        #     sheet,
        #     yearly_data.num_incarcerations,
        #     yearly_col_idx,
        #     incarcerations_row,
        # )
        # try_set_metric_cell(
        #     sheet,
        #     yearly_data.num_absconsions,
        #     yearly_col_idx,
        #     absconsions_row,
        # )


def write_metrics_from_sandbox(
    sheet: Worksheet,
    officer_id: str,
    officer_aggregated_metrics: OfficerAggregatedMetricsFromSandbox,
) -> None:
    """Write aggregated metrics to the appropriate cells in the sheet"""
    early_discharge_row = get_row_index(RowHeadingEnum.EARLY_TERMINATION_GRANTS)
    timely_risk_assessment_row = get_row_index(RowHeadingEnum.TIMELY_RISK_ASSESSMENTS)
    timely_contacts_row = get_row_index(RowHeadingEnum.TIMELY_CONTACTS)
    caseload_type_row = get_row_index(RowHeadingEnum.CASELOAD_TYPE)

    for metric in officer_aggregated_metrics.monthly_data[officer_id]:
        parsed_date = (metric.end_date_exclusive - relativedelta(days=1)).strftime(
            _COLUMN_DATE_FORMAT
        )
        col_idx = get_column_index(parsed_date)
        try_set_metric_cell(
            sheet, metric.task_completions_early_discharge, col_idx, early_discharge_row
        )
        try_set_metric_cell(
            sheet,
            metric.timely_risk_assessment,
            col_idx,
            timely_risk_assessment_row,
            FORMAT_PERCENTAGE_00,
        )
        try_set_metric_cell(
            sheet,
            metric.timely_contact,
            col_idx,
            timely_contacts_row,
            FORMAT_PERCENTAGE_00,
        )
        caseload_type_cell = sheet[f"{col_idx}{caseload_type_row}"]
        caseload_type_cell.value = metric.caseload_type
        caseload_type_cell.alignment = Alignment(wrap_text=True)

    yearly_col_idx = get_column_index(ColumnHeadingEnum.YEARLY.value)
    if officer_id in officer_aggregated_metrics.yearly_data:
        yearly_data = officer_aggregated_metrics.yearly_data[officer_id]
        try_set_metric_cell(
            sheet,
            yearly_data.task_completions_early_discharge,
            yearly_col_idx,
            early_discharge_row,
        )
        try_set_metric_cell(
            sheet,
            yearly_data.timely_risk_assessment,
            yearly_col_idx,
            timely_risk_assessment_row,
            FORMAT_PERCENTAGE_00,
        )
        try_set_metric_cell(
            sheet,
            yearly_data.timely_contact,
            yearly_col_idx,
            timely_contacts_row,
            FORMAT_PERCENTAGE_00,
        )
        caseload_type_cell = sheet[f"{yearly_col_idx}{caseload_type_row}"]
        caseload_type_cell.value = yearly_data.caseload_type
        caseload_type_cell.alignment = Alignment(wrap_text=True)


def write_impact_funnel_metrics(
    sheet: Worksheet, officer_id: str, impact_funnel_metrics: ImpactFunnelMetrics
) -> None:
    """Write impact funnel metrics to the appropriate cells in the sheet"""
    task_type_to_row_headings = {
        "EARLY_DISCHARGE": (
            RowHeadingEnum.EARLY_TERMINATION_ELIGIBLE_VIEWED,
            RowHeadingEnum.EARLY_TERMINATION_ELIGIBLE_NOT_VIEWED,
            RowHeadingEnum.EARLY_TERMINATION_ELIGIBLE_NOT_VIEWED_30_DAYS,
            RowHeadingEnum.EARLY_TERMINATION_MARKED_INELIGIBLE,
        ),
    }

    for metric in impact_funnel_metrics.data[officer_id]:
        parsed_date = (metric.date - relativedelta(days=1)).strftime(
            _COLUMN_DATE_FORMAT
        )
        col_idx = get_column_index(parsed_date)
        (
            eligible_viewed_row,
            eligible_not_viewed_row,
            eligible_not_viewed_30_days_row,
            marked_ineligible_row,
        ) = task_type_to_row_headings[metric.task_type]
        try_set_metric_cell(
            sheet, metric.eligible_viewed, col_idx, get_row_index(eligible_viewed_row)
        )
        try_set_metric_cell(
            sheet,
            metric.eligible_not_viewed,
            col_idx,
            get_row_index(eligible_not_viewed_row),
        )
        try_set_metric_cell(
            sheet,
            metric.eligible_not_viewed_30_days,
            col_idx,
            get_row_index(eligible_not_viewed_30_days_row),
        )
        try_set_metric_cell(
            sheet,
            metric.marked_ineligible,
            col_idx,
            get_row_index(marked_ineligible_row),
        )
        if metric.date == date(2025, 1, 1):
            try_set_metric_cell(
                sheet,
                metric.eligible_viewed,
                get_column_index(ColumnHeadingEnum.YEARLY.value),
                get_row_index(eligible_viewed_row),
            )
            try_set_metric_cell(
                sheet,
                metric.eligible_not_viewed,
                get_column_index(ColumnHeadingEnum.YEARLY.value),
                get_row_index(eligible_not_viewed_row),
            )
            try_set_metric_cell(
                sheet,
                metric.eligible_not_viewed_30_days,
                get_column_index(ColumnHeadingEnum.YEARLY.value),
                get_row_index(eligible_not_viewed_30_days_row),
            )
            try_set_metric_cell(
                sheet,
                metric.marked_ineligible,
                get_column_index(ColumnHeadingEnum.YEARLY.value),
                get_row_index(marked_ineligible_row),
            )


def write_grant_rate_formulas(sheet: Worksheet) -> None:
    """Write formulas to calculate the monthly grant rate / average monthly grant rate"""
    avg_caseload_row_idx = get_row_index(RowHeadingEnum.AVG_DAILY_CASELOAD)
    yearly_col_idx = get_column_index(ColumnHeadingEnum.YEARLY.value)
    for month in range(1, 16):
        col_idx = chr(ord(yearly_col_idx) + month)
        for grant_rate_row, num_grants_row in [
            (
                RowHeadingEnum.EARLY_TERMINATION_MONTHLY_GRANT_RATE,
                RowHeadingEnum.EARLY_TERMINATION_GRANTS,
            ),
        ]:
            grant_rate_row_idx = get_row_index(grant_rate_row)
            num_grants_row_idx = get_row_index(num_grants_row)
            cell = sheet[f"{col_idx}{grant_rate_row_idx}"]
            # Grant rate = # grants / avg caseload. If avg caseload is empty, set grant rate to empty.
            # Excel formula example: =IF(C8, C17/C8, "")
            cell.value = f'=IF({col_idx}{avg_caseload_row_idx}, {col_idx}{num_grants_row_idx}/{col_idx}{avg_caseload_row_idx}, "")'
            cell.number_format = FORMAT_PERCENTAGE_00

    first_date_col = chr(ord(yearly_col_idx) + 1)
    last_date_col = chr(ord(yearly_col_idx) + 15)
    for grant_rate_row in [
        RowHeadingEnum.EARLY_TERMINATION_MONTHLY_GRANT_RATE,
    ]:
        grant_rate_row_idx = get_row_index(grant_rate_row)
        cell = sheet[f"{yearly_col_idx}{grant_rate_row_idx}"]
        # Yearly column for grant rate is the average grant rate for each month in the year with data
        cell.value = f'=IFERROR(AVERAGE({first_date_col}{grant_rate_row_idx}:{last_date_col}{grant_rate_row_idx}), "")'
        cell.number_format = FORMAT_PERCENTAGE_00


def write_logins(sheet: Worksheet, officer_id: str, logins: Logins) -> None:
    num_logins_row = get_row_index(RowHeadingEnum.NUM_LOGINS)
    for month, num_logins in logins.data[officer_id].items():
        parsed_date = month.strftime(_COLUMN_DATE_FORMAT)
        col_idx = get_column_index(parsed_date)
        cell = sheet[f"{col_idx}{num_logins_row}"]
        cell.value = num_logins

    # Sum the data to get # logins in the year
    year_col_idx = get_column_index(ColumnHeadingEnum.YEARLY.value)
    first_date_col = chr(ord(year_col_idx) + 1)
    last_date_col = chr(ord(year_col_idx) + 12)
    sheet[
        f"{year_col_idx}{num_logins_row}"
    ] = f"=SUM({first_date_col}{num_logins_row}:{last_date_col}{num_logins_row})"

    # Count number of months where they logged in
    num_months_logged_in_row = get_row_index(RowHeadingEnum.NUM_MONTHS_LOGGED_IN)
    sheet[
        f"{year_col_idx}{num_months_logged_in_row}"
    ] = f'=COUNTIF({first_date_col}{num_logins_row}:{last_date_col}{num_logins_row}, ">0")'


def write_snooze_metrics(
    sheet: Worksheet, officer_id: str, snooze_metrics: SnoozeMetrics
) -> None:
    opportunity_type_to_row_heading = {
        "earlyTermination": RowHeadingEnum.EARLY_TERMINATION_MARKED_INELIGIBLE_REASON,
    }

    yearly_col_idx = get_column_index(ColumnHeadingEnum.YEARLY.value)
    metric = snooze_metrics.data[officer_id]
    for (
        opportunity_type,
        ineligible_reason_row,
    ) in opportunity_type_to_row_heading.items():
        cell = sheet[f"{yearly_col_idx}{get_row_index(ineligible_reason_row)}"]
        cell.value = metric.get(opportunity_type, "N/A")


def generate_sheets(aggregated_metrics_sandbox_prefix: str) -> None:
    """Read data and generate spreadsheets"""
    output_dir = os.path.join(os.path.dirname(__file__), "output")
    if not os.path.exists(output_dir):
        os.mkdir(output_dir)

    bq_client = BigQueryClientImpl()
    supervisors_to_officers = SupervisorsAndOfficers.from_bigquery(bq_client).data
    officer_aggregated_metrics = OfficerAggregatedMetrics.from_bigquery(bq_client)
    officer_aggregated_metrics_from_sandbox = (
        OfficerAggregatedMetricsFromSandbox.from_bigquery(
            bq_client, aggregated_metrics_sandbox_prefix
        )
    )
    impact_funnel_metrics = ImpactFunnelMetrics.from_bigquery(bq_client)
    logins = Logins.from_bigquery(bq_client)
    snooze_metrics = SnoozeMetrics.from_bigquery(bq_client)
    for supervisor, officers in supervisors_to_officers.items():
        wb = Workbook()
        for officer in officers:
            sheet = wb.create_sheet(title=f"{officer.officer_name.formatted()}")
            create_headers(
                sheet,
                officer.officer_name.formatted(),
                officer.officer_email,
                officer.officer_district,
                supervisor.supervisor_name.formatted(),
            )
            apply_conditional_formatting(sheet)
            write_metrics(sheet, officer.officer_id, officer_aggregated_metrics)
            write_metrics_from_sandbox(
                sheet, officer.officer_id, officer_aggregated_metrics_from_sandbox
            )
            write_impact_funnel_metrics(
                sheet, officer.officer_id, impact_funnel_metrics
            )
            write_grant_rate_formulas(sheet)
            write_logins(sheet, officer.officer_id, logins)
            write_snooze_metrics(sheet, officer.officer_id, snooze_metrics)

        # Remove the default sheet, since we created new sheets for our data.
        wb.remove(wb.worksheets[0])
        wb.save(
            f"{output_dir}/{supervisor.supervisor_district} - {supervisor.supervisor_name.formatted()}.xlsx"
        )


def parse_arguments(argv: List[str]) -> Tuple[argparse.Namespace, List[str]]:
    """Parses the required arguments."""
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--aggregated_metrics_sandbox_prefix",
        help="""Prefix of aggregated_metrics data loaded to a sandbox, for use in metric
        calculations not provided by our standard materialized views.
        See branch danawillow/id-perf-sandbox for the changes required to be loaded.""",
        type=str,
        required=True,
    )
    return parser.parse_known_args(argv)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    args, _ = parse_arguments(sys.argv)
    with local_project_id_override(GCP_PROJECT_PRODUCTION):
        generate_sheets(args.aggregated_metrics_sandbox_prefix)
