# Recidiviz - a data platform for criminal justice reform
# Copyright (C) 2024 Recidiviz, Inc.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
# =============================================================================
"""
Script for generating Excel spreadsheets with staff metrics for supervisors in ID, for use in their
performance reviews. See https://www.notion.so/recidiviz/Design-Doc-ID-2025-Performance-Reviews-1517889f4d198077b276fc1fc5569165
for details.

Usage: python -m recidiviz.tools.insights.generate_perf_review_spreadsheets
This will generate spreadsheets in the root directory of pulse-data.
"""

import logging
from enum import Enum
from random import random

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from openpyxl.worksheet.worksheet import Worksheet


class RowHeadingEnum(Enum):
    TIMELY_RISK_ASSESSMENTS = "Timely Risk Assessments"
    TIMELY_F2F_CONTACTS = "Timely F2F Contacts"
    SUPERVISION_LEVEL_MISMATCH = "Supervision & Risk Level Mismatch"


class ColumnHeadingEnum(Enum):
    YEARLY = "All of 2024"


_ROW_HEADINGS = [
    "Usage",
    "# Total Logins each month",
    "# of Months in 2024 where they logged in at least once",
    "Caseload Type",
    None,
    "Outcomes Metrics",
    "# Average Daily Caseload",
    "# Absconsions",
    "Months flagged as having a high absconsion rate",
    "# Incarcerations",
    "Months flagged as having a high incarceration rate",
    None,
    "Early Discharge",
    "Eligible & Viewed as of end of time period",
    "Eligible & Not Viewed as of end of time period",
    "Overridden (Marked Ineligible) as of end of time period",
    "Grants during time period",
    "Monthly grant rate",
    'Most often used "Ineligible Reason"',
    None,
    "LSU",
    "Eligible & Viewed as of end of time period",
    "Eligible & Not Viewed as of end of time period",
    "Overridden (Marked Ineligible) as of end of time period",
    "Grants during time period",
    "2024 average of monthly grant rate",
    'Most often used "Ineligible Reason"',
    None,
    "Supervision Level Downgrade",
    "Eligible & Viewed as of end of time period",
    "Eligible & Not Viewed as of end of time period",
    "Overridden (Marked Ineligible) as of end of time period",
    "Grants during time period",
    "2024 average of monthly grant rate",
    'Most often used "Ineligible Reason"',
    None,
    "Past FTRD",
    "Eligible & Viewed as of end of time period",
    "Eligible & Not Viewed as of end of time period",
    "Overridden (Marked Ineligible) as of end of time period",
    "Grants during time period",
    "Monthly grant rate",
    'Most often used "Ineligible Reason"',
    None,
    "Operations Metrics",
    RowHeadingEnum.TIMELY_RISK_ASSESSMENTS.value,
    RowHeadingEnum.TIMELY_F2F_CONTACTS.value,
    RowHeadingEnum.SUPERVISION_LEVEL_MISMATCH.value,
]

_COLUMN_HEADINGS = [
    None,
    ColumnHeadingEnum.YEARLY.value,
    "Jan 2024",
    "Feb 2024",
    "Mar 2024",
    "Apr 2024",
    "May 2024",
    "Jun 2024",
    "Jul 2024",
    "Aug 2024",
    "Sep 2024",
    "Oct 2024",
    "Nov 2024",
    "Dec 2024",
]

# Row numbers (1-indexed) that represent a section header
_ROW_SECTION_HEADER_INDICES = [2, 7, 14, 22, 30, 38, 46]

_GRAY_BACKGROUND = PatternFill(
    start_color="d9d9d9", end_color="d9d9d9", fill_type="solid"
)
_RED_BACKGROUND = PatternFill(
    start_color="ea9999", end_color="ea9999", fill_type="solid"
)
_BOLD_FONT = Font(bold=True)


CONDITIONAL_FORMAT_THRESHOLDS = {
    RowHeadingEnum.TIMELY_RISK_ASSESSMENTS: 0.9,
    RowHeadingEnum.TIMELY_F2F_CONTACTS: 0.9,
    RowHeadingEnum.SUPERVISION_LEVEL_MISMATCH: 0.98,
}


def create_headers(sheet: Worksheet) -> None:
    """Adds header rows + columns to the spreadsheet. This must be run before other data is added to
    the spreadsheet, because it just appends full rows."""
    data = [
        _COLUMN_HEADINGS,
        *[[row] for row in _ROW_HEADINGS],
    ]
    for row in data:
        sheet.append(row)

    sheet.column_dimensions["A"].width = 48
    for index in _ROW_SECTION_HEADER_INDICES:
        sheet[f"A{index}"].font = _BOLD_FONT
        sheet[f"A{index}"].fill = _GRAY_BACKGROUND
    sheet["B1"].font = _BOLD_FONT


def get_row_index(heading: RowHeadingEnum) -> int:
    return (
        _ROW_HEADINGS.index(heading.value)
        + 2  # 1-indexed + additional 1 for empty first row
    )


def get_column_index(heading: ColumnHeadingEnum) -> str:
    return chr(ord("A") + _COLUMN_HEADINGS.index(heading.value))


def apply_conditional_formatting(sheet: Worksheet) -> None:
    for row_heading, threshold in CONDITIONAL_FORMAT_THRESHOLDS.items():
        row_idx = get_row_index(row_heading)
        col_idx = get_column_index(ColumnHeadingEnum.YEARLY)

        cell = f"{col_idx}{row_idx}"

        sheet.conditional_formatting.add(
            cell,
            FormulaRule(
                formula=[f"AND(NOT(ISBLANK({cell})), {cell}<{threshold})"],
                fill=_RED_BACKGROUND,
            ),
        )


def write_random_data(sheet: Worksheet) -> None:
    for row in range(2, 50):
        # We don't need to put data in the section header rows or the rows before them
        if row in _ROW_SECTION_HEADER_INDICES or row + 1 in _ROW_SECTION_HEADER_INDICES:
            continue
        for col in range(ord("B"), ord("O")):
            sheet[f"{chr(col)}{row}"] = random()


def generate_sheets() -> None:
    wb = Workbook()
    # TODO(#35928): Create one spreadsheet per supervisor with one sheet per officer
    for i in range(0, 3):
        sheet = wb.create_sheet()
        sheet.title = f"Sheet {i}"

        create_headers(sheet)
        apply_conditional_formatting(sheet)
        # TODO(#35926): Write real data instead of random data
        write_random_data(sheet)

        # Format numerical data as a percentage
        for row in range(2, 50):
            for col in range(ord("B"), ord("O")):
                sheet[f"{chr(col)}{row}"].number_format = FORMAT_PERCENTAGE

    # Remove the default sheet, since we created new sheets for our data.
    wb.remove(wb.worksheets[0])
    wb.save("test_spreadsheet.xlsx")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)

    generate_sheets()
